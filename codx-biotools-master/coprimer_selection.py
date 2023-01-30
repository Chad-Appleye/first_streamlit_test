import csv
import pickle
import os
from re import TEMPLATE
import pandas as pd
import numpy as np
import itertools as itt

from openpyxl import load_workbook
from oligotools import coprimers_to_dataframe, binding_template


def delete_dir(dir):
    for f in os.listdir(dir):
        os.remove(dir+'/'+f)
    os.rmdir(dir)


class Selection:

    def __init__(self, file_path, regression_path, classification_path, template_sequences=None):
        """
        Class to handle the processing and prediction of CoPrimers performance from the CoPrimer design software.

        Params: 
            file_path: Path to the .xlsx workbook with the CoPrimer designs. Each target should be in a different sheet of the workbook
            regression_path: path to the saved regression model created by the Training class in training.py file
            classification_path: path to the saved classification model created by the Trianing class in training.py
            template_sequences (lst, optional) -- list of sequences of the DNA template for each target to find the binding template sequence. If no sequence given for a target, leave as None at that index (default None)

        Output: 
            None: use methods to created the desired output 
        """
        self.class_filename = classification_path
        self.reg_filename = regression_path
        self.path = file_path
        self.targets = self._sheet_names()
        if template_sequences is None:
            template_sequences = [None] * len(self.targets)
        else:
            self.templates = template_sequences
        #models
        self.rf, self.rf_class = self._load_models()


    def predict(self, master_mix="MME", n_outputs=2, write_csvs=False, output_dir=None): 
        """ main function for coprimer predictions
        Params:

            master_mix: the type of master mix that will be used for the assay, MME or BHQ
            n_outputs: the number of forwards and reverses you want as an ouput, default=2
            write_csvs: (Optional) Write the predictions from each target to a CSV file
            output_dir: (Optional) path to directory where output files will be written. Default will create "predictions" folder in current working directory
            


        Output: return data frame with predictions, excel file with best predictions will be written 
        """
        if output_dir is None:
            output_dir = './predictions'

        best_primers_df = pd.DataFrame()

        #loop through all the sheets
        print('Calculating...')
        sequence_dict = {}
        for sheet_name, template_sequence in zip(self.targets, self.templates):
            print(sheet_name)

            forwards, reverses = self._split_forwards_reverses(sheet_name)

            #new data for predictions 
            predict_df = self._combinations(master_mix)

            #data frame with predictions 
            out_df, pred, pred_pd = self._make_predictions(predict_df, sheet_name)

            if write_csvs == True:
                #send the df to a csv file in case of further review needed
                
                os.makedirs(output_dir, exist_ok=True)

                out_df.to_csv(os.path.join(output_dir, f'{sheet_name}.csv'), index=False)

            #redefine out_df with only top 10%
            #also return list of the best unique forward and reverse primers
            out_df, f_primers, r_primers = self._filter(out_df, n_outputs, pred, pred_pd)


            if template_sequence is not None:
                template_sequence = template_sequence.strip().upper()
                template_sequence = [b for b in template_sequence if b in ['A', 'T', 'G', 'C', 'U']]
                template_sequence = ''.join(template_sequence)
            else:
                template_sequence = ''
            
            sequence_dict.update({sheet_name: template_sequence})

            #get final data frame sorted by predictor variable with penalty for primer dimer formation
            best_primers_df = best_primers_df.append(self._predictor_variable(out_df, f_primers, r_primers, sheet_name, forwards, reverses, n_outputs))

        
        best_primers_df['template_sequence'] = best_primers_df['Target'].map(sequence_dict)
        best_primers_df.index += 1
        os.makedirs(output_dir, exist_ok=True)
        best_primers_df['is_reverse'] = best_primers_df['OligoName'].str.contains(r'\.R\d+$', regex=True)
        best_primers_df['binding_sequence'] = best_primers_df.apply(lambda x: binding_template(x.Sequence, x.template_sequence, x.is_reverse, x.Gap), axis=1) 
        best_primers_df.drop(columns=['is_reverse', 'template_sequence'], inplace=True)
        print('Finished')
        best_primers_df.to_excel(os.path.join(output_dir, 'CoPrimer_predictions.xlsx'), index=False)
        print(best_primers_df.head())

        # remove folders and files used in calculation 
        delete_dir('CSVs')
        

    def _load_models(self):
        #import saved models
        rf = pickle.load(open(self.reg_filename, 'rb'))
        rf_class = pickle.load(open(self.class_filename, 'rb'))

        return rf, rf_class
    

    def _sheet_names(self):
        #read Sheets
        wb = load_workbook(filename = self.path)
        return wb.sheetnames

    
    def _split_forwards_reverses(self, sheet_name):
        """ Use regex to identify forwards and reverse CoPrimers and split them up into their own csv files, the files will be          overwritten for each loop
        :param path: the path to the excel workbook with new CoPrimer designs to use in predictions
        :param sheet_name: current sheet name in the loop
        
        :return: two data frames, one for forwards, the other for reverses. Two csv files will be written so they can be read later
        """
        input_df = coprimers_to_dataframe(self.path, sheet_name=sheet_name) #create pandas dataframe
        forwards = input_df[input_df['OligoName'].str.contains(r'\.F\d+$', regex=True)] #use regex to determine forward CoPrimers
        reverses = input_df[input_df['OligoName'].str.contains(r'\.R\d+$', regex=True)] #determine reverse CoPrimers

        # make folder for csvs 
        csv_folder = 'CSVs/'
        os.makedirs(os.path.dirname(csv_folder), exist_ok=True)
        
        #create csv of both dataframes
        forwards.to_csv(r'CSVs/Forwards.csv', index=False, header=False)
        reverses.to_csv(r'CSVs/Reverses.csv', index=False, header=False)

        return forwards, reverses 


    def _combinations(self, master_mix):
        """Make all combinations possible from the primers given
        :param master mix: the master mix that will be used in the assay MME or BHQ

        :returns: data frame containing all possible combinations and the attributes
        """
        forwards_path = r"CSVs/Forwards.csv" #path to the forwards dataframe created in self.split_forwards_reverses
        reverses_path = r"CSVs/Reverses.csv" #path to reverses dataframe
    
        header = '(f)Sequence,(f)OligoName,(f)Pos,(f)Length,(f)PrimerLen,(f)CaptureLen,(f)Gap,(f)PrimerTm,(f)CaptureTm,(f)CombinedTM,(f)TeP,(f)PL,(f)TotalDiff,(f)SC,(f)PrimerEFF,(f)CombinedEFF,(f)PSC,(f)PHC,(f)PGC,(f)CSC,(f)CHC,(f)CGC,(r)Sequence,(r)OligoName,(r)Pos,(r)Length,(r)PrimerLen,(r)CaptureLen,(r)Gap,(r)PrimerTm,(r)CaptureTm,(r)CombinedTM,(r)TeP,(r)PL,(r)TotalDiff,(r)SC,(r)PrimerEFF,(r)CombinedEFF,(r)PSC,(r)PHC,(r)PGC,(r)CSC,(r)CHC,(r)CGC'
        
        # create directory for the combinations matrix
        combinations_folder = 'CSVs/'
        os.makedirs(os.path.dirname(combinations_folder), exist_ok=True)

        with open(forwards_path, newline='', encoding='utf-8-sig') as f_1, open(reverses_path, newline='', encoding='utf-8-sig') as f_2:
            reader_1 = csv.reader(f_1)
            reader_2 = csv.reader(f_2)
            combs = itt.product(reader_1, reader_2)
            rows_gen = (l_1 + l_2 for l_1, l_2 in combs)

            with open('CSVs/Combinations Matrix.csv', 'w', newline='') as out_file:
                writer = csv.writer(out_file)
                writer.writerow(header.split(','))
                writer.writerows(rows_gen)


        #import new values to predict in pandas df
        predict_df = pd.read_csv(r"CSVs/Combinations Matrix.csv", encoding='utf-8-sig')
        predict_df = predict_df.drop(columns=['(f)Sequence', '(f)OligoName', '(r)Sequence', '(r)OligoName', '(f)TotalDiff', '(r)TotalDiff'])

        if master_mix == 'MME':
            predict_df['Master Mix_MME'] = 1
            predict_df['Master Mix_BHQ'] = 0

        elif master_mix == 'BHQ':
            predict_df['Master Mix_MME'] = 0
            predict_df['Master Mix_BHQ'] = 1
        
        return predict_df


    def _make_predictions(self, predict_df, sheet_name):
        """Make predictions for CT value and PD classification on new data
        :param predict_df: a dataframe with new CoPrimers to predict on from combinations()
        :param sheet_name: the current sheet in the loop

        :returns: out_df, a dataframe with additional rows for the prediction values
        """

        pred = self.rf.predict(predict_df) #Regression predictions
        pred_pd = self.rf_class.predict(predict_df) #primer dimer predictions
        
        
        out_df = pd.read_csv(r"CSVs/Combinations Matrix.csv", usecols=['(f)OligoName', '(r)OligoName'])
        out_df['CT Prediction'] = np.round(pred, decimals=2)
        out_df['PD Prediction'] = pred_pd
        out_df = out_df.sort_values(by='CT Prediction')

        return out_df, pred, pred_pd


    def _filter(self, out_df, n_outputs, pred, pred_pd):
        """filter values from predictions so selections on made only on the top 10% of CoPrimers

        :param out_df: output from predict(), the predicted values included in the data frame 
        :param pred: regression predictions
        :param pred_pd: primer dimer predictions

        :return: shortened data frame with top 10% of CoPrimers
        """  

        #keep only top 10% of the predicted primers if there are more than 50 predictions
        if len(out_df) > 20:
            out_df = out_df[:len(out_df)//10]
        
        #list the primers in the 10% of predictions
        f_primers = out_df['(f)OligoName'].unique().tolist()
        r_primers = out_df['(r)OligoName'].unique().tolist()
        
        #make sure there are enough unique values for the output required
        n=9
        while len(f_primers) < n_outputs or len(r_primers) < n_outputs:
            if n == 0:
                print("Not enough unique predictions in the data. Select fewer outputs")
                break
            print('Not enough unique values, trying again')
            out_df = pd.read_csv(r"CSVs/Combinations Matrix.csv", usecols=['(f)OligoName', '(r)OligoName'])
            out_df['CT Prediction'] = np.round(pred, decimals=2)
            out_df['PD Prediction'] = pred_pd
            out_df = out_df.sort_values(by='CT Prediction')
            out_df = out_df[:len(out_df)//n]
            
            f_primers = out_df['(f)OligoName'].unique().tolist()
            r_primers = out_df['(r)OligoName'].unique().tolist()
            
            n -= 1        
        
        return out_df, f_primers, r_primers


    def _predictor_variable(self, out_df, f_primers, r_primers, sheet_name, forwards, reverses, n_outputs):
        """calculate new predictor variable that takes primer dimer formation as penalty to CoPrimer selection
        :params out_df: shortened data frame with top performing Coprimer pairs
        :param f_primers: list of all forward primers, output from split_forwards_reverses()
        :param r_primers: list of all reverese primers, ouput from split_forwards_revereses()
        :param sheet_name: current sheet in loop
        :param forwards: data frame with only forward CoPrimers from split_forwards_revereses()
        :param reverses: data frame with only reverse CoPrimers from split_forwards_revereses()
        :param n_outputs: specified number of requested outputs

        :returns: new data frame sorted by lowest score of predictor variable
        """
        #df with only forward or reverse CoPrimers
        forwards_df = out_df.drop(columns='(r)OligoName')
        reverses_df = out_df.drop(columns='(f)OligoName')
        
        
        #Calculate the percentage of primer dimer predictions per CoPrimer
        f_dicts = []
        for primer in f_primers:
            f_mask = forwards_df[forwards_df['(f)OligoName']==primer]
            pd_percent = (sum(f_mask['PD Prediction'])/len(f_mask))*100
            pd_percent = np.round(pd_percent, decimals=3)
            ave_ct = round(f_mask['CT Prediction'].mean(), 3)
            f_dict = {'OligoName': primer, 'CT Average': ave_ct, 'PD Percent': pd_percent}
            f_dicts.append(f_dict)

            
        forwards_df = pd.DataFrame(f_dicts) #redefine the forwards dataframe
        r_dicts = []
        for primer in r_primers:
            r_mask = reverses_df[reverses_df['(r)OligoName']==primer]
            pd_percent = (sum(r_mask['PD Prediction'])/len(r_mask))*100
            pd_percent = np.round(pd_percent, decimals=3)
            ave_ct = round(r_mask['CT Prediction'].mean(), 3)
            r_dict = {'OligoName': primer, 'CT Average': ave_ct, 'PD Percent': pd_percent}
            r_dicts.append(r_dict)
            
        
        reverses_df = pd.DataFrame(r_dicts) #redefine reverses DataFrame
        

        # Calculate predictor value = CT + %PD/50
        forwards_df['Predictor'] = round(forwards_df['CT Average'] + forwards_df['PD Percent']/50, 3)
        reverses_df['Predictor'] = round(reverses_df['CT Average'] + reverses_df['PD Percent']/50, 3)
                
        forwards_df = forwards_df.sort_values(by='Predictor', ascending=True)
        reverses_df = reverses_df.sort_values(by='Predictor', ascending=True)
        


        #define best CoPrimers in lists
        best_f = [i for i in forwards_df['OligoName'].unique()][:n_outputs]
        best_r = [i for i in reverses_df['OligoName'].unique()][:n_outputs]
    
        #set indices as oligo names to make indexing easier in next step
        forwards_index = forwards_df.set_index('OligoName') 
        reverses_index = reverses_df.set_index('OligoName')
        

        best_primers = []
        for i in best_f:

            params_dict = dict(Target = sheet_name)
            index_f = forwards.index[forwards['OligoName']==i][0]
            f_mask = out_df[out_df['(f)OligoName']==i] #mask the out dataframe to find the primer in question
            pd_percent = forwards_index.loc[i]['PD Percent']
            ave_ct = forwards_index.loc[i]['CT Average']
            predictor = forwards_index.loc[i]['Predictor']
            params_dict.update({'Predictor Value': predictor, 'Average CT': ave_ct, '% Primer Dimer':pd_percent})
            params_dict.update(forwards.loc[index_f].to_dict())
            best_primers.append(params_dict)
            
        for i in best_r:  
            params_dict_2 = dict(Target = sheet_name)
            r_mask = out_df[out_df['(r)OligoName']==i] #Mask reverses
            index_r = reverses.index[reverses['OligoName']==i][0]
            r_pd_percent = reverses_index.loc[i]['PD Percent']
            r_ave_ct = reverses_index.loc[i]['CT Average']
            r_predictor = reverses_index.loc[i]['Predictor']
            params_dict_2.update({'Predictor Value': r_predictor, 'Average CT': r_ave_ct, '% Primer Dimer':r_pd_percent})
            params_dict_2.update(reverses.loc[index_r].to_dict())
            best_primers.append(params_dict_2)
            

        return best_primers
    