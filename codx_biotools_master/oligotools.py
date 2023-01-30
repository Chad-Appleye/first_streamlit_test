import numpy as np
import Bio
import os
import re
import pandas as pd
import glob
import subprocess as sp
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import random
import itertools as itt

from Bio import Seq, SeqIO, Entrez, SearchIO
from Bio.Blast import NCBIWWW
from Bio.SeqUtils import GC
# from tools import replace_all, slidingWindow, remove_dir, get_sheet_names, rand_subset, parent_dir, flatten
from itertools import groupby, product
from copy import deepcopy
from Bio.pairwise2 import align, format_alignment
from Bio.SeqIO.FastaIO import SimpleFastaParser
from tqdm import tqdm
from numpy.random import default_rng


# ----------------------- global variables----------------------------------------------

# keywords used to detect the presence of dyes or quenchers in oligo sequences
dyes_and_quenchers = ['QUASAR', 'FAM', 'BHQ', 'CAL', 'FLUOR', 'CAL', 'FLUOR', 'RED', 'ORANGE', 'YELLOW', 'GREEN', 
    'QUENCH', 'QUENCHER', 'TAMRA', 'MGB', 'TET']

ambiguous_bases = ['M', 'R', 'W', 'S', 'Y', 'K', 'V', 'H', 'D', 'B', 'N']


# --------------------------- functions --------------------------------------------------

def calculate_tm(sequence):
    """Calculates the Tm of an oligo according to University of Utah Tm Tool,
    https://www.dna.utah.edu/tm/SW_Methods_Tm_Tool.pdf"""
    sequence = sequence.upper()
    concentration = 0.15e-6
    mono_salts = 50e-3
    free_mg = 5e-3
    deltaH = {'AA': -7.9, 'AT': -7.2, 'TA': -7.2, 'CA': -8.5, 'GT': -8.4, 
              'CT': -7.8, 'GA': -8.2, 'CG': -10.6, 'GC': -9.8, 'GG': -8.0, 
              'TT': -7.9, 'TG': -8.5, 'AC': -8.4, 'AG': -7.8, 'TC': -8.2, 
              'CC': -8.0, 'init_gc': 0.1, 'init_at': 2.3}
    deltaS = {'AA': -22.2, 'AT': -20.4, 'TA': -21.3, 'CA': -22.7, 'GT': -22.4, 
              'CT': -21.0, 'GA': -22.2, 'CG': -27.2, 'GC': -24.4, 'GG': -19.9, 
              'TT': -22.2, 'TG': -22.7, 'AC': -22.4, 'AG': -21.0, 'TC': -22.2,
              'CC': -19.9, 'init_gc': -2.8, 'init_at': 4.1}
    sum_deltaH = 0
    sum_deltaS = 0
    for pairs in np.arange(len(sequence)):
        pair = sequence[pairs:pairs + 2]
        if len(pair) == 2:
            sum_deltaH += deltaH[pair]
            sum_deltaS += deltaS[pair]
    if sequence[:2] == 'GC' or 'CG':
        sum_deltaH += deltaH['init_gc']
        sum_deltaS += deltaS['init_gc']
    if sequence[:2] == 'AT' or 'TA':
        sum_deltaH += deltaH['init_at']
        sum_deltaS += deltaS['init_at']
    salt_correction = 0.368 * (len(sequence) - 1) * np.log(mono_salts + 3.795 *
                                                           np.sqrt(free_mg))
    tm = ((sum_deltaH * 1000) / \
         (sum_deltaS + salt_correction + 1.9872 * np.log(concentration - \
         (concentration / 2)))) - 273.15
    return tm


def reverse_complement(sequence):
    """Calculates the reverse complement of a DNA sequence"""
    sequence_object = Seq.Seq(sequence)
    rev_complement = str(sequence_object.reverse_complement())
    if "[" in rev_complement or "]" in rev_complement:
        rev_complement = rev_complement.replace('[', '1')
        rev_complement = rev_complement.replace(']', '2')
        rev_complement = rev_complement.replace('1', ']')
        rev_complement = rev_complement.replace('2', '[')
    return rev_complement


def complement(sequence):
    """Calculates the complement of a DNA sequence"""
    sequence_object = Seq.Seq(sequence)
    complement = str(sequence_object.complement())
    if "[" in complement or "]" in complement:
        complement = complement.replace('[', '1')
        complement = complement.replace(']', '2')
        complement = complement.replace('1', ']')
        complement = complement.replace('2', '[')
    return complement


def calculate_gc(sequence):
    """Calculcates the percent GC content of a DNA sequence"""
    return GC(sequence)


def convert_snps(sequence):
    """Converts SNP format from '[X/Y]' to the corresponding IUPAC 
    representation"""
    iupac_dict = {'[A/G]': 'R', '[G/A]': 'R', '[C/T]': 'Y', '[T/C]': 'Y',
                  '[A/C]': 'M', '[C/A]': 'M', '[G/T]': 'K', '[T/G]': 'K',
                  '[C/G]': 'S', '[G/C]': 'S', '[A/T]': 'W', '[T/A]': 'W',
                  '[A/*]': 'A', '[T/*]': 'T', '[C/*]': 'C', '[G/*]': 'G'}
    new_sequence = replace_all(sequence, iupac_dict)
    return new_sequence


def reduce_ambiguous(sequence):
    """Replaces IUPAC ambiguous base code with one allele represented
    by that code. E.g. R is replaced by A, Y is replaced by C."""
    iupac_dict = {'R': 'A', 'Y': 'C', 'M': 'A', 'K': 'G', 'S': 'C', 'W': 'T'}
    new_sequence = replace_all(sequence, iupac_dict)
    return new_sequence


def score_sequence(sequence, window_size):
    """Provides a weighted score of a sequence for primer placement based on 
    GC content, degeneracies, and nucleotide repeates. 
    Weights --
    GC content: 1;
    Degeneracies: 1.5;
    Nucleotide repeats 5 bases and greater: 1.5
    """
    g_c_weight = 1
    degen_weight = 1.5
    repeat_weight = 1.5
    if len(sequence) < window_size:
        window_size = len(sequence)
    window = slidingWindow(sequence, window_size)
    total_score_list = []
    for window_slice in window:
        slice_seq = window_slice._data
        slice_no_degens = slice_seq.translate(
            {ord(c): None for c in 'RYSWKMBDHVN'})
        optimal_g_c = 60
        g_c = 10 * (slice_no_degens.count('g') +
                    slice_no_degens.count('c')) / len(slice_no_degens)
        g_c_score = abs(g_c - optimal_g_c)
        degen_score = 2 * np.sum(np.array([slice_seq.count(n) for n in 
                                           'RYSWKM'])) + \
            3 * np.sum(np.array([slice_seq.count(n) for n in 'BDHV'])) + \
            4 * slice_seq.count('N')
        repeats_list = find_repeats(slice_seq)
        repeats_score = np.sum(
            np.array([len(repeat) for repeat in repeats_list]))
        total_score = np.average([g_c_score, degen_score, repeats_score],
                                 weights=[g_c_weight, degen_weight, 
                                          repeat_weight])
        total_score_list.append(total_score)
    return np.ravel(np.array(total_score_list))


def find_repeats(sequence, min_repeat_length=4):
    """Identifies single-nucleotide repeats longer than the specified
    min_repeat_length."""
    repeats = []
    for _, group in groupby(sequence):
        group = ''.join(group)
        if len(group) > min_repeat_length:
            repeats.append(group)
    return repeats


def expand_ambiguous_dna(seq, ignore_n=False):
    """Returns a list of all possible sequences given an ambiguous DNA input. 
    SNPs must be represented in IUPAC format, e.g. Y for C/T change."""
    degenerate_dict = deepcopy(Bio.Data.IUPACData.ambiguous_dna_values)
    if ignore_n:
        degenerate_dict.update({'N': 'N'})
    return list(map("".join, product(*map(degenerate_dict.get, seq))))


def random_oligo(length, gc, dna=True):
    """Generate a random oligo with a specified length and GC content.
    Args:
        length: The length of the oligo.
        gc: The GC content of the oligo(0 - 100).
        dna (bool, Optional) -- if True, will return DNA. Will return RNA if False (Default True)
    Output:
        template_string: An oligo with A, T, C, and G.
    """
    g_c_count = int(length * (0.01 * gc))
    a_t_count = length - g_c_count
    # self.np.random.seed(seed)
    # seed_state = self.np.random.get_state()
    # self.np.random.set_state(seed_state)
    g_c = np.random.randint(0, 2, g_c_count).astype(str)
    # self.np.random.set_state(seed_state)
    a_t = np.random.randint(0, 2, a_t_count).astype(str)
    np.place(g_c, g_c == '0', 'G')
    np.place(g_c, g_c == '1', 'C')
    np.place(a_t, a_t == '0', 'A')
    np.place(a_t, a_t == '1', 'T')
    template_string = np.concatenate((g_c, a_t))
    np.random.shuffle(template_string)
    sequence = ''.join(template_string)

    if dna is False:
        sequence = sequence.replace('T', 'U')
    return sequence


def split_forwards_reverses(data, sheet_name=0, write_csv=False):
        """ Use regex to identify forwards and reverse CoPrimers and split them up into their own csv files, the files will be          overwritten for each loop
        :param data: the path to the excel workbook or pandas dataframe with new CoPrimer designs to use in predictions
        :param sheet_name: specified sheet name if reading xlsx file
        
        :return: forwards, reverses: two data frames, one for forwards, the other for reverses
        """
        if isinstance(data, str):
            if data.endswith('.csv') or data.endswith('.txt'):
                input_df = pd.read_csv(data)
            elif data.endswith('.xlsx'):
                input_df = coprimers_to_dataframe(data, sheet_name=sheet_name)
            else:
                raise Exception('not a valid file path')
        elif type(data) == pd.core.frame.DataFrame:
            input_df = data
        else:
            raise Exception(f'{type(data)} is not a valid data type')
        forwards = input_df[input_df['OligoName'].str.contains(r'\.F\d+$', regex=True)] #use regex to determine forward CoPrimers
        reverses = input_df[input_df['OligoName'].str.contains(r'\.R\d+$', regex=True)] #determine reverse CoPrimers
        
        if write_csv: 
            #put the csvs into a folder in the current working directory
            csv_folder = './CSVs'
            if not os.path.isdir(csv_folder):
                os.mkdir(csv_folder)

            #create csv of both dataframes
            forwards.to_csv(r'CSVs/Forwards.csv', index=False, header=False)
            reverses.to_csv(r'CSVs/Reverses.csv', index=False, header=False)

        return forwards, reverses 


def trim_template(sequence, length=120):
    """ Trim a sequence positive control template to a specified size from the middle to keep binding regions on both ends
    Parameters:
        sequence (str) -- the sequence to trim
        length (int, optional) -- the final length of the sequence to return (Default 120)
    Return:
        sequence (str) -- trimmed sequence with bases removed from the middle until it reaches the specified length
    """
    try:
        length = int(length)
    except:
        raise Exception('The length must be an integer')
    
    seq_length = len(sequence)

    if seq_length <= length:
        raise Exception('the sequence length is already {}'.format(seq_length))

    sequence = list(sequence)
    while len(sequence) > length:
        sequence.pop(len(sequence)//2)

    return ''.join(sequence)


def reverse_transcribe(sequence):
    """ Convert an RNA sequence to DNA. This will also remove lower-case 'r' values and 5'-3' notation if they exist in the sequnce 
    Arguments:
        sequence (str) -- Any RNA sequence
    Return: 
        DNA sequence (str)
    """

    sequence = sequence.strip().replace('r', '').upper().replace('U', 'T')

    return ''.join(i for i in sequence if i.isalpha())


def transcribe(sequence, order_format=False):
    """ Convert DNA sequence to RNA. Will remove 5' and 3' notations and any non-alpha characters
    Arugments:
        sequence (str) -- Any DNA sequence
        order_format (bool, Optional) -- if True, will format the RNA sequence for ordering through suppliers with a small 'r' before each base
    Return 
        RNA sequence (str)
    """

    sequence = sequence.strip().upper().replace('T', 'U')

    if order_format is True:
        return ''.join('r' + i for i in sequence if i.isalpha())
    else:
        return ''.join(i for i in sequence if i.isalpha())


def bind_primer(primer, sequence, file_name=None, method='local'):
    """ 
    Uses Bio.pairwise2 and optimized binding scores for CoPrimers to align a primer with a sequence to see the best binding location. The output will be printed to the console
    Arguments:
        primer (str) -- primer sequence to test
        sequence (str) -- sequence that the primer will bind to
        file_name (str, optional) -- path where file will be written. If None, no file will be written (Default None)
        method (str, optional) -- 'local' or 'global' for the method to use in showing binding region. If the sequence is long, use local. (Default local)
    Return
        None. binding will be shown in console
    """
    method = method.strip().lower()

    if method == 'local':
        alignment = align.localms(primer, sequence, 2, -1, -2, -.5)
    elif method == 'global':
        alignment = align.globalms(primer, sequence, 2, -1, -2, -.5)
    else:
        raise Exception(f'{method} is not a valid option for method. Please choose either local or global')
    
    for a in alignment:
        print(format_alignment(*a))
    
    if not file_name is None:
        with open(file_name, 'w+') as f:
            for a in alignment:
                f.write(format_alignment(*a))


def is_exact_match(primer, sequence, check_reverse_complement=True):
    """ 
    Check if the primer is an exact match to the sequence using regex
    Arguments:
        primer (str) -- linearized primer sequence with 'N' in place of each gap. If reverse, should be the reverse compliment
        sequence (str) -- Sequence of the binding region or genome to test if the primer binds exactly
        check_reverse_complement (bool) -- if True, will check if the reverse complement of the primer matches if the given sequence does not
    Return 
        Bool If exact match found, returns True, else returns False
    """
    primer = primer.strip().upper()
    sequence = sequence.strip().upper()

    pattern = re.compile(primer.replace('N', '.'))
    match = pattern.search(sequence)
    if match:
        return True
    else:
        if check_reverse_complement == True:
            primer = reverse_complement(primer)
            pattern = re.compile(primer.replace('N', '.'))
            match = pattern.search(sequence)
            if match:
                print('found reverse complement')
                return True
            else:
                return False
        else:
            return False


def coprimers_to_dataframe(coprimers, sheet_name=0):
    """ Read an .xlsx excel workbook containing CoPrimer designs into a dataframe. The function will remove all data outside of the CoPrimer table, and will only keep the table
    of CoPrimers from the file.  
    Arguments:
        coprimers (str) -- Path to CoPrimer design excel workbook 
        sheetname (str, list, int, None) -- specify which sheet(s) to read into the dataframe see sheet_name argument for pandas.read_excel()
    Return: 
        pandas DataFrame of CoPrimers 
    """
    # if there are multiple sheets to concatenate together, recursively call this function to format and join them together
    if sheet_name is None or isinstance(sheet_name, list):  
        if sheet_name is None: # if no sheet name specified, get all sheet names
            sheet_name = get_sheet_names(coprimers)
        dfs = []
        for i in sheet_name:
            dfs.append(coprimers_to_dataframe(coprimers, sheet_name=i))
        return pd.concat(dfs)


    df = pd.read_excel(coprimers, engine='openpyxl', sheet_name=sheet_name)
    

    cols = df.columns
    if 'OligoName' not in cols and '(f)OligoName' not in cols and '(r)OligoName' not in cols:
        df = pd.read_excel(coprimers, engine='openpyxl', sheet_name=sheet_name)


    df.dropna(axis=1, how='all', inplace=True)
    df.dropna(axis=0, how='any', inplace=True)
    df.drop_duplicates(inplace=True)

    cols = df.columns
    if 'OligoName' not in cols and '(f)OligoName' not in cols and '(r)OligoName' not in cols:
        df.columns = df.iloc[0]
        df.drop(df.index[0], inplace=True)
    df.reset_index(inplace=True, drop=True)

    return df


def find_gap_length(coprimer, template, reverse=False):
    """ 
    find the gap length between primer and capture sequences of coprimers. 
    
    coprimer input must be linearized with at least one "N" representing the gap between the capture and primer
    
    Arguments:
        coprimer (str) -- Sequence of coprimer with the linker denoted with "N" or "[]" between the capture and primer
        template (str) -- Template sequence where the coprimer will bind. Can be a DNA sequence of any length, not just the binding region
        reverse (bool) -- If the CoPrimer is a reverse primer pass True, else pass False (default False)
    Returns:
        int number of gap bases 
    """
    template = template.strip().upper()
    coprimer = coprimer.strip().upper()
    split_primer = coprimer.split('N')
    if len(split_primer) <= 1:
        coprimer = coprimer.replace(']', '')
        split_primer = coprimer.split('[')

    primer = split_primer[0]
    capture = split_primer[-1]

    for i in range(1, 10):
        coprimer = capture + 'N' * i + primer
        if reverse == True:
            coprimer = reverse_complement(coprimer)
        if is_exact_match(coprimer, template):
            return i
    print('CoPrimer gap length cannot be determined. Check the sequences')
    return 0


def linearize_coprimer(coprimer, reverse=False, gap=None):
    """ linearizes the coprimer so it is in the order it will bind to the DNA template
    
    This will return only one "N" in the place of the gap between the priming and capture sequences unless gap is specified. Use find_gap_length to get the correct number of gaps for the CoPrimer 
    Arguments:
        coprimer (str) -- the sequnece of the CoPrimer to be linearized in the format output from the CoPrimer design software
        reverse (bool, optional) -- whether the CoPrimer is a reverse CoPrimer or not (default False)
        gap (int, optional) -- the gap length between the primer and capture, if None, will insert a single "N" for the gap
    Returns:
        (str) -- sequence of the linearized coprimer with one "N" to represent the gap between the primer and capture
    """
    
    coprimer = clean_sequence(coprimer)
    primer = coprimer.split('[')[0]
    capture = coprimer.split(']')[-1]
    if gap != None:
        try:
            int(gap)
        except:
            raise Exception('gap must be an integer value')
        
        coprimer = capture + 'N' * gap + primer

    else:
        coprimer = capture + 'N' + primer 
        
    if reverse is True:
        coprimer = reverse_complement(coprimer)
    return coprimer


def binding_template(coprimer, template, reverse=False, gap=None):
    """ Finds the sequence of the template where the coprimer binds -- the binding sequence will only be found if the CoPrimer matches the template exactly
    Arguments:
        coprimer (str) -- The sequence of the CoPrimer as output from the CoPrimer design software
        template (str) -- the sequence of the template where the coprimer binds
        reverse (bool, optional) -- whether the CoPrimer is a reverse CoPrimer or not (default False)
        gap (int, optional) -- the gap length between the primer and capture. If None, the gap will be found from the template (default None)
    Returns:
        (str) -- the sequence of the template binding region
    """
    template = template.strip().upper()
    coprimer = coprimer.strip().upper()
    primer = coprimer.split('[')[0]
    capture = coprimer.split(']')[-1]

    if gap != None:
        try:
            int(gap)
        except:
            raise Exception('gap must be an integer')
        gap_length = gap
    else:
        coprimer = linearize_coprimer(coprimer, reverse=reverse, gap=gap)
        gap_length = find_gap_length(coprimer, template)
    
    
        
    
    coprimer = capture + 'N' * gap_length + primer
    

    if reverse is True:
        coprimer = reverse_complement(coprimer)
        

    pattern = re.compile(coprimer.replace('N', '.'))
    match = pattern.search(template)
    if match:
        span = match.span()
        return template[span[0]: span[1]]
    else:
        return ''


def fasta_stats(fasta_path):
    """ 
    print the stats about the number of sequences and length of the sequences

    Arguments:
        fasta_path (str) -- Path to the fasta file
    Return:
        None -- stats will be printed to the console
    """
  
    df = fasta_to_df(fasta_path)
    df['sequence_length'] = df['sequence'].apply(lambda x: len(x))

    print(df['sequence_length'].describe())


def fasta_to_df(fasta_path):
    """ 
    read a FASTA file into a dataframe with 2 columns name, and sequence
    Arguments:
        fasta_path (str) -- path to the FASTA file 
    Return:
        pandas DataFrame
    """
    records = []
    with open(fasta_path, 'r') as f:
        for name, seq in SimpleFastaParser(f):
            tmp = {
                'name': '>' + name, 
                'sequence': seq.strip().upper()
            }
            records.append(tmp)
    
    return pd.DataFrame.from_dict(records)


def translate(oligo_sequence):
    table = {
        'ATA':'I', 'ATC':'I', 'ATT':'I', 'ATG':'M',
        'ACA':'T', 'ACC':'T', 'ACG':'T', 'ACT':'T',
        'AAC':'N', 'AAT':'N', 'AAA':'K', 'AAG':'K',
        'AGC':'S', 'AGT':'S', 'AGA':'R', 'AGG':'R',                 
        'CTA':'L', 'CTC':'L', 'CTG':'L', 'CTT':'L',
        'CCA':'P', 'CCC':'P', 'CCG':'P', 'CCT':'P',
        'CAC':'H', 'CAT':'H', 'CAA':'Q', 'CAG':'Q',
        'CGA':'R', 'CGC':'R', 'CGG':'R', 'CGT':'R',
        'GTA':'V', 'GTC':'V', 'GTG':'V', 'GTT':'V',
        'GCA':'A', 'GCC':'A', 'GCG':'A', 'GCT':'A',
        'GAC':'D', 'GAT':'D', 'GAA':'E', 'GAG':'E',
        'GGA':'G', 'GGC':'G', 'GGG':'G', 'GGT':'G',
        'TCA':'S', 'TCC':'S', 'TCG':'S', 'TCT':'S',
        'TTC':'F', 'TTT':'F', 'TTA':'L', 'TTG':'L',
        'TAC':'Y', 'TAT':'Y', 'TAA':'_', 'TAG':'_',
        'TGC':'C', 'TGT':'C', 'TGA':'_', 'TGG':'W',
    }

    while len(oligo_sequence) % 3 != 0:
        oligo_sequence = oligo_sequence[:-1]
    
    protein = ""

    for i in range(0, len(oligo_sequence), 3):
        codon = oligo_sequence[i: i + 3]
        protein += table[codon]
    
    return protein


def multi_sequence_align(sequences_path, reference_path, out_dir, email):
    """ 
    use command-line tool ViralMSA.py to run multi-sequence alignment. ViralMSA.py must be in path as well as Minimap2
    Arguments:
        sequences_path (str) -- path to the multi-sequence fasta to use in the alignment
        reference_path (str) -- path to the single-sequence fasta to use as a reference
        out_dir (str) -- output directory where alignment results will be written. (Cannot be existing directory)
        email (str) -- email address to pass to ViralMSA.py 

    """
    if os.path.isdir(out_dir):
        remove_dir(out_dir)
    if not os.path.isfile(sequences_path):
        raise Exception('Sequences path does not exist')
    if not os.path.isfile(reference_path):
        raise Exception('reference path does not exist')
    
    cmd = f'ViralMSA.py -s {sequences_path} -r {reference_path} -e {email} -o {out_dir}'
    sp.check_output(cmd, shell=True)


def clean_sequence(seq, allow_ambiguous=False, special_characters=None):
    """ remove unwanted characters from a sequence 
    Arguments:
        seq (str) -- sequence input to clean
        allow_ambiguous (bool) -- True will allow all IUPAC ambiguous bases in the final sequence, False will remove all ambiguous (Default False)
        special_characters (iterable) -- Special characters to allow in the output sequence. (Default None)
    Return:
        sequence (str) -- sequence that has been formatted according to specifications
    
    """
    seq = seq.upper()
    replacements = dyes_and_quenchers

    if special_characters != None:
        try:
            special_characters = list(special_characters)
        except:
            raise TypeError('special characters must be passed in an iterable like a list or a tuple')

    for i in replacements:
        seq = seq.replace(i, '')
    seq = seq.replace(' ', '') # remove spaces
    
    if seq.count('[]') == 1 or seq.count('[]') == 2: #check if it's already a cleaned CoPrimer or CopTer sequence
        pass
    else:
        seq = seq.replace('[', '').replace(']', '')
        seq = re.sub(r'SP[A-Z]*\s*\d+', ' ', seq) # Replace Spacer 18 with [] to remove the linker in CoPrimers
        if ' ' in seq:
            seq = re.split(r'\s+', seq)
            if len(seq) == 3:
                chars = ['G', 'C', 'T'] # allow for polyAs in the linker, and differentiate between CoPrimers and Copters
                if any(char in seq[1] for char in chars):
                    seq = '[]'.join(seq)
                else:
                    del seq[1]
                    seq = '[]'.join(seq)
            else:
                seq = '[]'.join(seq)


    allowable_characters = ['A', 'T', 'G', 'C', 'U', '[', ']', 'I'] # baseline allowable characters
    if allow_ambiguous is True:
        allowable_characters = allowable_characters + ambiguous_bases
    if special_characters:
        allowable_characters = allowable_characters + special_characters

    seq = [i if i in allowable_characters else '' for i in seq]
    return ''.join(seq)


def trim_to_tm(sequence, tm, method, return_tm=False, quiet=False):
    """ Trim an oligo sequence close to a desired melting temperature
    Arguments:
        sequence (str) -- sequence to trim
        tm (int, float) -- desired melting temperature of the oligo
        method ['left', 'right', 'middle'] -- method to trim the oligo. left will trim from the left, right will trim from the right, and middle will trim from the middle character
        return_tm (bool) -- if True, will return the sequence and the TM in a dictionary (Default False)
        quiet (bool) -- if True, print statements will be silenced
    return: 
        trimmed oligo
    """

    #clean the sequence
    sequence = clean_sequence(sequence)
    # make sure the tm is an integer value
    try:
        tm = float(tm)
    except:
        raise Exception('tm must be a numeric value')

    # allow the tm to be .5 degrees above the specified tm
    tm = tm + 0.5

    # check that an appropriate value for method has been passed
    method = method.strip().lower()
    methods = ['left', 'right', 'middle']
    if method not in methods:
        raise Exception(f'method must be one of: {methods}')
    
    actual_tm = calculate_tm(sequence)
    
    if actual_tm < tm:
        if quiet != True:
            print(f'calculated TM of {actual_tm} already below desired {tm}')
        if return_tm is True:
            return {
                'sequence': sequence,
                'tm': actual_tm
            }
        else:
            return sequence

    while calculate_tm(sequence) > tm:
        if method == 'left':
            sequence = sequence[1:]
        if method == 'right':
            sequence = sequence[:-1]
        if method == 'middle':
            sequence_list = list(sequence)
            sequence_list.pop(len(sequence)//2)
            sequence = ''.join(sequence_list)
    
    if return_tm is True:
        return {
            'sequence': sequence, 
            'tm': calculate_tm(sequence)
        }

    else:
        return sequence

    
def df_to_fasta(df, out_file):
    """ 
    write a fasta file from a sequence dataframe
    Arguments:
        df (DataFrame) -- pandas dataframe with name and sequence columns
        out_file (path, str) -- output file path for the fasta file to be written
    return:
        None - file will be written
    """
    
    #error handling
    if not out_file.lower().endswith('.fasta') and not out_file.endswith('.aln'):
        out_file = out_file + '.fasta'

    out_dir = parent_dir(out_file)
    if out_dir != '':
        if not os.path.isdir(out_dir):
            os.makedirs(out_dir, exist_ok=True)

    columns = list(df)
    if 'name' not in columns or 'sequence' not in columns:
        raise Exception('name and sequence columns not found in DataFrame')

    # make sure there's a '>' sign at the beginning of the name column value
    df['name'] = df['name'].apply(lambda x: '>' + x if x[0] != '>' else x)
    df['sequence'] = df['sequence'].apply(lambda x: x.upper())

    df = df[['name', 'sequence']]

    df.to_csv(out_file, index=False, header=False, sep='\n')
    return 


def combine_fastas(out_path, *fasta_paths):
    """ 
    Takes any number FASTA files and combine those into a single multi-FASTA file
    Arguments
        out_path (str) -- path where output multi-FASTA file will be written
        fasta_paths (str) -- path to multiple FASTA files
    Return
        None -- FASTA file will be written
    """

    # create iterable of dataframes to concatenate
    dfs = []
    for fasta in fasta_paths:
        dfs.append(fasta_to_df(fasta))
    # concatenated dataframe
    df = pd.concat(dfs)
    df_to_fasta(df, out_path) # send to new FASTA file
    return


def get_seq_from_fasta(fasta_path, index=None):
    """ 
    Get a sequence from a fasta file. Quickly extract a sequence from a single FASTA, or a specific sequence from a multi-FASTA
    Arguments:
        fasta_path (str) -- path to the FASTA file
        index (int) -- index of the sequence desired, if None the first sequence will be returned
    Return:
        sequence (str) -- the sequence from the FASTA file
    """

    if index != None:
        try:
            int(index)
        except:
            raise TypeError('index must be an integer value')

    else:
        index = 0


    i = 0
    for seq in SeqIO.parse(fasta_path, 'fasta'):
        if i == index:
            return str(seq.seq)
        else: 
            continue
        i += 1

    print('Index out of range')
    return None


def binding_location(primer, reference, mismatches=0, method='first', return_if_reverse=False):
    """ 
    Find where a primer or oligo binds on a reference seqeunce

    Arguments: 
        primer (str) -- oligo sequence to bind to reference, if using a CoPrimer - must use linearized CoPrimer sequence
        reference (str) -- sequence or path to reference sequence 
        mismatches (int) -- number of mismatches allowed in the binding (0-3) Default=0
        method (str) -- method of binding location - one of ['first', 'span']

    Return:
        binding_location (int) -- integer value where the primer or oligo will begin to bind
    """

    # check if reference is passed as a path
    if os.path.isfile(reference):
        reference = get_seq_from_fasta(reference)

    method = method.strip().lower()    
    if method not in ['first', 'span']:
        raise Exception(f'method must be either first or span, {method} passed')

    primer = primer.strip().upper()
    reference = reference.strip().upper()
    all_primers = mismatch_combinations(primer, mismatches)

    for primer in all_primers:
        pattern = re.compile(primer.replace('N', '.').replace('I', '.'))
        match = pattern.search(reference)
        if match:
            if method == 'first':
                if return_if_reverse == True:
                    return match.span()[0], False
                else:
                    return match.span()[0]
            elif method == 'span':
                if return_if_reverse == True:
                    return match.span(), False
                else:
                    return match.span()
        else:
            pattern = re.compile(reverse_complement(primer).replace('N', '.').replace('I', '.'))
            match = pattern.search(reference)
            if match:
                print('found the reverse complement')
                if method == 'first':
                    if return_if_reverse == True:
                        return match.span()[0], True
                    else:
                        return match.span()[0]

                elif method == 'span':
                    if return_if_reverse == True:
                        return match.span(), True
                    else:
                        return match.span()
            else:
                continue
    else:
        raise Exception('No binding found for that oligo and reference')


def coprimer_pair_evaluation(forward, reverse, alignment, normalize=False):
    """ 
    Evaluate the number of sequences in the alignment that will be amplified with no mismatches on either the forward or reverse. Will give the number (or proportion) of sequences in 
    the alignment that have no mismatches on the forward and no mismatches on the reverse

    Arguments:
        forward (str, dict) -- forward CoPrimer sequence - already linearized CoPrimer with "N"s separating the capture and primer
            if a dictionary of CoPrimer names and sequences is passed, either CoPrimer will be used in the calculation. This is to evaluate designs where 
            multiple CoPrimers could amplify the same locus to improve coverage. 
        reverse (str, iterable) -- reverse CoPrimer - already linearized CoPrimer with "N"s separating the capture and primer
            if a dictionary of CoPrimer names and sequences is passed, either CoPrimer will be used in the calculation. This is to evaluate designs where 
            multiple CoPrimers could amplify the same locus to improve coverage.
        alignment (str) -- path to the alignment file for the sequence data
        normalize (bool) -- if True, the return will be the percent of sequences with exact matches for the forward and reverse CoPrimers
    Return:
        number of sequences (or proportion) with exact matches for the forward and reverse CoPrimers
    """

    # read in the sequence data as a dataframe
    df = fasta_to_df(alignment)
    print(f'calculating mismatches for forward and reverse on {len(df)} sequences')

    if type(forward) is str:
        multiple_forwards = False 
        df['forward'] = df['sequence'].apply(lambda x: is_exact_match(forward, x))
    elif type(forward) is dict: # do calculation for each forward if multiple are passed
        multiple_forwards = True
        for name, seq in forward.items():
            df[name] = df['sequence'].apply(lambda x: is_exact_match(seq, x))
    else:
        raise TypeError('type str or dict required for forward argument')
    
    # repeat process for reverse
    if type(reverse) is str:
        multiple_reverses = False
        df['reverse'] = df['sequence'].apply(lambda x: is_exact_match(reverse, x))
    elif type(reverse) is dict: 
        multiple_reverses = True
        for name, seq in reverse.items():
            df[name] = df['sequence'].apply(lambda x: is_exact_match(seq, x))
    else:
        raise TypeError('type str or dict required for reverse argument')

    # if multiple forwards or reverses are passed, then a true for any forward and any reverse satisfies the condition
    def both_true(x):
        if multiple_forwards:
            forward_condition = False # initiate condition to False
            for f in forward.keys():
                if x[f]: 
                    forward_condition = True
                    break
                else:
                    continue
            #if no true found, the condition is not met
            if forward_condition is False:
                return False

        else:
            if x['forward']:
                forward_condition = True
            else:
                return False
        
        # check if reverse condition is met
        if multiple_reverses:
            reverse_condition = False # initiate condition to False
            for r in reverse.keys():
                if x[r]:
                    reverse_condition = True
                    break
                else:
                    continue
            if reverse_condition is False:
                return False # if no Reverse, the 'AND' condition is not met
        else:
            if x['reverse']:
                reverse_condition = True
            else:
                return False
        return True # if this line is reached, then a forward and reverse match was found
    

    # use both_true to now evaluate if the forward "AND" reverse condition is met
    df['both_true'] = df.apply(lambda x: both_true(x), axis=1)

    total = df['both_true'].sum()

    if normalize is True:
        return total / len(df)
    else:
        return total

    
def linearize_all(coprimers_df, sheet_name=0):
    """ 
    Split the forward and reverse CoPrimers in a dataframe, linearize them and return the dataframe with a 'linear' column of linearized CoPrimers

    The dataframe must contain columns "Sequence", "OligoName", and "Gap"
    Arguments:
        coprimers_df (DataFrame) -- DataFrame or path to coprimers workbook 
        sheet_name (str, list, None) -- specify the sheet name to read if giving an excel workbook with multiple sheets 

    return:
        coprimers_df (DataFrame) -- DataFrame with added "linear" column
    """

    df = coprimers_df
    if type(df) != pd.core.frame.DataFrame:
        if os.path.isfile(df):
            df = coprimers_to_dataframe(df, sheet_name=sheet_name)
        else:
            raise Exception('Not a valid DataFrame or file to CoPrimers workbook')

    cols = df.columns
    
    # make the column names not case-sensitive
    if 'Sequence' not in cols:
        raise Exception('Sequence was not found in the columns')
    elif 'OligoName' not in cols:
        raise Exception('OligoName was not found in the columns')
    elif 'Gap' not in cols:
        raise Exception('Gap not found in the columns')


    forwards, reverses = split_forwards_reverses(df)

    forwards['linear'] = forwards.apply(lambda x: linearize_coprimer(x['Sequence'], reverse=False, gap=x['Gap']), axis=1)
    reverses['linear'] = reverses.apply(lambda x: linearize_coprimer(x['Sequence'], reverse=True, gap=x['Gap']), axis=1)

    return pd.concat([forwards, reverses])
    

def gc_graph(sequence, window):
    """ 
    return a graph of the rolling G-C percent of a given sequence
    Arguments:
        sequence (str) -- Sequence to analyze or path to a single FASTA file
        window (int) -- Size of the rolling window for the G-C percentage to graph
    
    Output:
        plotly figure object 
    """

    # error handling 
    try:
        window = int(window)
    except ValueError:
        raise Exception('window must be an integer value')
    
    # extract sequence if path given
    if os.path.isfile(sequence):
        sequence = get_seq_from_fasta(sequence)

    # clean the sequence
    sequence = clean_sequence(sequence, allow_ambiguous=True)

    # calculate the rolling GC content
    def rolling_gc(window, df):
        rolling_gc_vals = np.empty(len(sequence))

        for i, b in enumerate(sequence):
            if i < window:
                rolling_gc_vals[i] = np.nan
            else:
                seq = sequence[i - window: i] # make the sequence go from current value to previous (window) length 
                rolling_gc_vals[i] = calculate_gc(seq)
        return rolling_gc_vals

    # construct DataFrame
    df = pd.DataFrame()
    df['position'] = range(1, len(sequence) + 1)
    df['rolling_GC'] = rolling_gc(window, df)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['position'], y=df['rolling_GC'], name=f'G-C percent - window={window}', showlegend=True))
    fig.update_layout(title='Rolling G-C percent')
    fig.update_yaxes(title_text='rolling G-C percent')
    return fig


def sequence_size_distribution(fasta_path, bins=None):
    """ 
    generates a histogram of the sequence sizes in a fastA file. It is the visual verson of fasta_stats

    Arguments:
        fasta_path (str) -- path to fastA file 
        bins (int) -- custom number of bins in the histogram (default None)
    
    Return:
        plotly graph object
    """

    # error handling 
    if not os.path.isfile(fasta_path):
        raise Exception('File not found')

    if bins != None:
        try:
            bins = int(bins)
        except:
            raise TypeError('bins must be an integer value')
    
    df = fasta_to_df(fasta_path)
    df['length'] = df['sequence'].apply(lambda x: len(x))
    fig = px.histogram(df, x='length', nbins=bins)
    fig.update_layout(title=f'Sequence Length Histogram -- {len(df)} sequences')
    fig.update_xaxes(title='sequence length')
    return fig


def get_input_sequences(excel_path, sheet_names=None):
    """ 
    if the input sequence for CoPrimer design is in the Excel workbook that will be used for selection, this method will extract that input sequence for each sheet_name and return them in a list

    This is useful for CoPrimer selections in finding the template sequences so that the binding sequence will be given in the output. The input sequence is identified by the preceding column being 'input'
    or by a sequence with ">>" present, which is standard notation for forward and reverse separation in CoPrimer design

    Arguments:
        excel_path (str) -- path to the excel sheet where the CoPrimers and input sequences can be found
        sheet_names (str, list) -- if passing a single sheet name, a string is accepted, otherwize a list or None is required. If None, all sheets will be used. (Default None)
    
    Return:
        list with the cleaned sequences present
    """

    if not os.path.isfile(excel_path):
        raise Exception(f'file {excel_path} found')
    wb = openpyxl.load_workbook(excel_path)

    seqs = []

    if sheet_names != None:
        if type(sheet_names) == str:
            sheet_names = [sheet_names]
        elif type(sheet_names) == list:
            pass
        else:
            raise Exception('sheet_names must be a string or a list')
    else:
        sheet_names = wb.sheetnames
    for sheet in sheet_names:
        active = wb[sheet]
        break_out_flag = False # break out of outer loop if sequence found
        for col in range(active.max_column):
            if break_out_flag == True:
                break
            for row in active.iter_rows():
                val = row[col].value
                try:
                    val = val.strip().lower()
                except:
                    pass
                if val == 'input':
                    seq = clean_sequence(row[col + 1].value)
                    seqs.append(seq)
                    break_out_flag = True
                    break
                elif '>>' in str(val):
                    seq = clean_sequence(row[col].value)
                    seqs.append(seq)
                    break_out_flag = True
                    break
    return seqs


def number_in_alignment(sequences, alignment, normalize=True):
    """ 
    Find the number of sequences in a FASTA file or dataframe that exist in an alignment. 
    Since the alignment file only outputs the acquisition number associated for each sequence, it's hard to know how many of a certain subset
    of sequences were successfully aligned. This function will compare the acquisition number from the FASTA file with the number in the alignment file
    to determine how many sequences appear in the alignment. 

    This is mostly useful for determining how many sequences of a subset were successfully aligned in a larger alignment 

    Arguments:
        sequences (str, DataFrame) -- path to FASTA file or dataframe from FASTA with the sequences in question
        alignment (str) -- path to the alignment .aln file output from ViralMSA to check how many sequences from the sequences argument
            appear in the alignment
        normalize (bool) -- if True, will return the proportion of sequences that appear in the alignment (Default True)

    Return:
        number of sequences that appear in the alignment (int)
    """
    if type(sequences) == str:
        if os.path.isfile(sequences):
            sequences = fasta_to_df(sequences)

        elif type(sequences) != pd.core.frame.DataFrame:
            raise Exception('sequences must be path to FASTA or DataFrame object')
    
    if not os.path.isfile(alignment):
        raise Exception('that alignement path does not exist')

    if 'name' not in sequences.columns:
        raise Exception("'name' not found in the sequences DataFrame columns")
    
    sequences['acquisition'] = sequences['name'].apply(lambda x: x.split(' ')[0])

    aln = fasta_to_df(alignment)
    names = aln['name']
    if normalize is False:
        return sum(sequences['acquisition'].isin(names))
    else:
        return round(sum(sequences['acquisition'].isin(names))/len(sequences), 2)


def get_reference(fasta_path, out_path, index=None):
    """
    Reads a multi-FASTA file and writes a sequence as a single FASTA for use as a reference.
    
    Arguments:
        fasta_path (str) -- path to the multi-FASTA input
        out_path (str) -- path to write the single FASTA reference
        index (int) -- index of record in FASTA file to write as reference. If None, the largest will be used. (Default None)
    Returns:
        None -- a single FASTA reference file will be written
    """

    if index == None:
        largest = ''
        
        with open(fasta_path, 'r') as f:
            for name, seq in SimpleFastaParser(f):
                if len(seq) > len(largest):
                    largest = seq
                    info = name

        with open(out_path, 'w+') as g:
            g.write('>' + info + '\n' + largest.upper())
    else:
        try:
            index = int(index)
        except ValueError:
            raise Exception('index must be an integer value')
        
        counter = -1
        file_written = False
        with open(fasta_path, 'r') as f:
            for name, seq in SimpleFastaParser(f):
                counter += 1
                if counter == index:
                    with open(out_path, 'w+') as g:
                        g.write('>' + name + '\n' + seq.upper())
                    file_written = True
                    break

                else:
                    continue
            if file_written is False:
                raise Exception(f'index {index} is out of range of the FASTA file')
            

def find_best_ref(fasta, email, max_iter=None):
    """ 
    Find the best reference sequence from a FASTA file by testing the alignment output from all sequences or a randomized subset. 
    Arguments:
        fasta (path) -- path to the multi-FASTA file of all sequneces
        max_iter (int) -- maximum number of iterations to try -- use to avoid long computation time for larger FASTA files (Default None)
        email (str) -- email to pass to the ViralMSA API

    Return:
        index of the sequence in the FASTA with highest success rate as a reference sequence
    """
    num_alignments = []
    all_indeces = []
    temp = './temp'
    ref = os.path.join(temp, 'ref.fasta')
    out = os.path.join(temp, 'alignment')
    os.makedirs(out, exist_ok=True)
    
    df = fasta_to_df(fasta)
    max_index = len(df)
    if max_iter is None:
        indeces = range(max_index)
        
    else:
        indeces = np.random.randint(max_index, size=max_iter)
    for i in tqdm(indeces):
        get_reference(fasta, ref, index=i)
        if os.path.isdir(out):
            remove_dir(out)
        multi_sequence_align(fasta, ref, out, email)
        aln = glob.glob(os.path.join(out, '*.aln'))[0]
        num_alignments.append(len(fasta_to_df(aln)))
        all_indeces.append(i)


    max_alignments = max(num_alignments)
    max_index = num_alignments.index(max_alignments)
    index = all_indeces[max_index]
    print(f'max alignments: {max_alignments}\nindex: {index}')
    remove_dir(temp)
    return index   


def get_template(primers, reference, overlap=1, min_length=None, max_length=None, mismatches=0):
    """ 
    Use the primers from the path to get a template sequence that has the binding location of all the primers

    Arguments:
        primers (list, str, DataFrame) -- List of primer sequences, CoPrimers DataFrame, or path to CoPrimers file -- CoPrimers must be linear and reverses must be the reverse complement 
        reference (str) -- reference sequence that contains binding location of the primers
        overlap (int) -- number of nucleotide bases to extend on each side of the binding location (Default=1)

    return: 
        template_sequence (str) -- template sequence that all primers will bind to
    """
    try:
        overlap = int(overlap)
    except:
        TypeError(f'overlap must be an integer, not {type(overlap)}')

    if max_length != None:
        try:
            max_length = int(max_length)
        except:
            TypeError(f'max_length must be integer, not {type(max_length)}')
    
    if min_length != None:
        try:
            min_length = int(min_length)
        except:
            TypeError(f'min_length must be integer, not {type(min_length)}')
    
    if min_length and max_length:
        if min_length > max_length:
            raise ValueError(f'min_length {min_length} is larger than max_length {max_length}')

    if isinstance(primers, dict): # allow a dictionary of primers with 'sequence' as a key
        primers = [val['sequence'] for val in primers.values()]


    if isinstance(primers, list):
        binding_starts = []
        binding_stops = []
        for primer in primers:
            try:
                start = binding_location(primer, reference, mismatches=mismatches)
            except:
                primer = reverse_complement(primer)
                start = binding_location(primer, reference, mismatches=mismatches)
            stop = start + len(primer)
            binding_starts.append(start)
            binding_stops.append(stop)

    elif isinstance(primers, pd.DataFrame):
        if 'linear' not in primers.columns:
            primers = linearize_all(primers)
        primers['start'] = primers['linear'].apply(lambda x: binding_location(x, reference))
        primers['stop'] = primers.apply(lambda x: x['start'] + len(x['linear']), axis=1)
        binding_starts = primers['start'].tolist()
        binding_stops = primers['stop'].tolist()
    
    elif isinstance(primers, pd.Series):
        return get_template(primers.tolist(), reference, overlap, min_length, max_length)

    else:
        raise TypeError(f'type {type(primers)} is not supported. Please pass an approved data type')

    min_binding = min(binding_starts)
    max_binding = max(binding_stops)
    
    adjusted_min = min_binding - overlap
    adjusted_max = max_binding + overlap

    ref_length = len(reference)

    if adjusted_min < 0:
        adjusted_min = 0
    
    if adjusted_max > ref_length:
        adjusted_max = ref_length


    template = reference[adjusted_min: adjusted_max]

    length = len(template)

    if max_length:
        if length > max_length:
            template = trim_template(template, length=max_length)
    
    if min_length:
        if min_length > ref_length:
            raise ValueError(f'minimum length {min_length} is larger than the reference length {ref_length}')
        while len(template) < min_length:
            if adjusted_min > 0:
                adjusted_min -= 1
            if adjusted_max < ref_length:
                adjusted_max += 1
            template = reference[adjusted_min: adjusted_max]
    

    bases = ['A', 'T', 'G', 'C', 'U']
    template = list(template)
    for i, b in enumerate(template):
        if b not in bases:
            template[i] = random.choice(bases[:-1])
        
    return ''.join(template)


def check_template(primer_names, primer_sequences, template):
    """ 
    Check if all primers bind to the template. 
    Arguments:
        primers_names (iterable) -- iterable object of primer names
        primer_sequences (iterable) -- iterable object of primer sequences - should correlate directly with the primer_names object
        template (seq) -- Template sequence or path to fasta file with the template sequence
    Return:
        If some primers don't bind to the template, those primers will be returned in a pandas DataFrame
    """

    names = primer_names
    seqs = primer_sequences
    
    if len(names) != len(seqs):
        raise Exception('primer_names and primer_sequences must be the same length')
    
    if os.path.isfile(template):
        template = get_seq_from_fasta(template)
    elif not isinstance(template, str):
        raise TypeError('template must be a sequence or a path to a FASTA file')
    
    df = pd.DataFrame({'name': names, 'seq': seqs})

    df['is_match'] = df['seq'].apply(lambda x: is_exact_match(x, template))

    binding = df[df["is_match"]==1]
    non_binding = df[df["is_match"]==0]
    if sum(df['is_match']) == len(primer_names):
        print('All primers bind to the template')
    elif sum(df['is_match']) == 0:
        print('None of the primers bind')
        return non_binding
    else:
        print(f'binding primers: {binding["name"].tolist()}')
        print(f'non-binding primers: {non_binding["name"].tolist()}')
        return non_binding.drop(columns=['is_match'])


def trim_aln(aln, out, start, end):
    """ 
    Trim every sequences in an alignment file to a specified region

    Arguments:
        aln (str, path) -- path to alignment file
        out (str, path) -- path to output alignment file
        start (int) -- starting index for trimmed sequence
        end (int) -- ending index for trimmed sequence
    """
    
    if not os.path.isfile(aln):
        raise FileExistsError(f'File {aln} not found')

    df = fasta_to_df(aln)
    df2 = df.copy(deep=True)
    df2['sequence'] = df['sequence'].apply(lambda x: x[start: end])
    
    df_to_fasta(df2, out)


def seq_matrix(fasta):
    """ 
    Create a numpy sequence matrix from an alignment FASTA file
    
    Arguments:
        fasta (str) -- path to a FASTA alignment file
    Return: 
        sequence_matrix (ndarray)

    """
    
    df = fasta_to_df(fasta)
    seqs = df['sequence']
    seq_mat = np.empty((len(df), len(df.iloc[0]['sequence'])), dtype=str)
    idx = 0
    for seq in tqdm(seqs):
        seq_mat[idx, :] = list(seq)
        idx += 1
    return seq_mat


def filter_fasta(fasta, out_path, min_length=100):
    """ 
    Filter out reads from a FASTA file that don't have sequence information. This is primarily useful if trimming an alignment file, then filtering out
    the reads that aren't sequenced in that region
    Arguments:
        fasta (str) -- path to FASTA file
        out_path (str) -- path to write output file
        min_length (int) -- minimum length of sequence to keep 
    return: None
    """

    if not os.path.isfile(fasta):
        raise FileExistsError(f'Invalid FASTA file path: {fasta}')

    records = []
    with open(fasta, 'r') as f:
        for name, seq in SimpleFastaParser(f):
            seq_mod = seq.lstrip('-')
            seq_mod = seq.rstrip('-')
            if len(seq) > min_length:
                tmp = {
                    'name': '>' + name, 
                    'sequence': seq.strip().upper()
                }        
                records.append(tmp)
            else:
                continue
    df = pd.DataFrame.from_dict(records)
    df_to_fasta(df, out_path)


def get_taxonomy(id, email):
    """ 
    Use sequence ID from NCBI to get a list of the taxonomical classifaction of that organism

    Arguments:
        id (str) -- ID number from NCBI 
        email (str) -- email to send to Entrez when querying the database
    
    return:
        (list) -- list of taxonomy from kindom [0] to species [-1]
    """
    Entrez.email = email
    record = Entrez.efetch(db='nucleotide', id=id, rettype='gb', retmode='text')
    info_handle = SeqIO.read(record, 'genbank')
    return info_handle.annotations['taxonomy']


def read_blast(file):
    """ 
    Read an XML file from an NCBI blast into a SearchIO object

    Arguments: 
        file (str) -- path to the file to be read
    Return:
        Bio.SearchIO object
    """

    # check if file exists and that it's an xml file
    if not os.path.isfile(file):
        raise FileExistsError(f'{file} does not appear to be a valid file path')
    if not file.lower().endswith('xml'):

        raise TypeError('file must be xml type')
    
    return Bio.SearchIO.read(file, 'blast-xml')


def blast(seq, file, entrez_query=None):
    """ 
    run a nucleotide blast through the NCBI webpage blast (requires internet connection). The blast results will be written in an XML format and parsed from the XML using Biopython

    Arguments: 
        seq (str) -- DNA sequence to run in the blast
        file (str) -- path where the XML output file will be written
        entrez_query (str) -- entrez query to limit search. See NCBI for more information

    Return : a Bio.SearchIO object with the blast results will be returned  
    """

    if not file.lower().endswith('xml'):
        file = file.replace('.', '') + '.xml'

    dirs_list = file.split('/')

    if len(dirs_list) > 1: # check if there's a directory to create
        dirs = '/'.join(dirs_list[:-1])

        os.makedirs(dirs, exist_ok=True) # allow the ability to specify directories that don't already exist
    
    
    kwargs = {
        'program': 'blastn', 
        'database': 'nt',
        'sequence': seq, 
        'gapcosts': "5 2", 
        'expect': 10, 
        'entrez_query': entrez_query, 
        'alignments': 5000, 
        'descriptions': 5000, 
        'hitlist_size': 5000, 
        'word_size': 11
    }
    result_handle = NCBIWWW.qblast(**kwargs)

    #write the output in an xml file
    with open(file, 'w') as f:
        f.write(result_handle.read())

    return read_blast(file)


def blast_report(blast_object, out_file, e_value=1, list_species=True, email=None, keyword_filters=None):
    """ 
    Generate a report of the blast to see the organisms with significant hits

    Arguments:
        blast_object (obj, str) -- the blast results read from the XML document or path to XML file
        out_file (str) -- path to output file to be written
        e_value (int) -- e value cutoff to include in the report
        list_species (bool) -- whether to list all unique species that appear in the report at the top
        email (str) -- if list_species is True, an email must be provided for entrez
        keyword_filters (list) -- a list of keywords to exclude from the report. If the hit description contains that keyword, it won't be added to the report
    Return: 
        None -- report file will be written 

    """

    if isinstance(blast_object, str): # read in the blast xml if file path is given
        if os.path.isfile(blast_object):
            blast_object = read_blast(blast_object)
        else:
            raise FileExistsError(f'{blast_object} is not a file')
    
    if list_species is True: # to find species, an email is required
        if email is None:
            raise Exception('You must provide an email to search species on Entrez')
        else:
            pass

    # some filters will be automatically applied 
    filters = ['synthetic', 'artificial']

    # add the user-defined keywords to the filters list
    if keyword_filters is not None:
        keyword_filters = [str(x).lower().strip() for x in keyword_filters]
        filters = filters + keyword_filters



    # creat sub directories to output file if necessary
    out_dirs = out_file.split('/')
    if len(out_dirs) > 1:
        out_dir = '/'.join(out_dirs[:-1])
        os.makedirs(out_dir, exist_ok=True)

    if not out_file.endswith('.txt'):
        out_file = out_file + '.txt'
    
    if not isinstance(e_value, int) or isinstance(e_value, float):
        try:
            e_value = float(e_value)
        except:
            raise TypeError(f'e_value must be numeric type, not {type(e_value)}')

    query_length = blast_object.seq_len # the length of the sequence passed to BLAST
    all_accessions = []
    hsps = []
    identities = []
    # iterate through the hits and the HSPs of the blast results object
    for hit in blast_object:
        # check if any of the filter keywords appear in the description 
        desc = hit.description.lower()
        if any(keyword in desc for keyword in filters):
            continue
        for hsp in hit:
            similarity = hsp.fragment.aln_annotation['similarity']
            hsp_len = sum(1 for i in similarity if i == '|')
            identity = hsp_len / query_length # identity is how much of the query length the hsp covers
        
            if hsp.evalue <= e_value and identity >= 0.8:
                if hit.accession not in all_accessions:
                    all_accessions.append(hit.accession)
                hsps.append(hsp)
                identities.append(f'{hsp_len}/{query_length}')

    if list_species is True:
        species = []
        for accession in all_accessions:
            spec = get_taxonomy(accession, email)
            if spec not in species:
                species.append(spec)
    
    # write the file
    with open(out_file, 'w') as f:
        if list_species is True:
            f.write('List of species present:\n')
            for spec in species:
                f.write(str(spec)+'\n')
            f.write('\n\n')
        if len(hsps) > 0:

            f.write('Highly Significant Pairs:' + '\n\n')
            for hsp, identity in zip(hsps, identities):
                f.write('\n\n' + str(hsp.hit_description) + '\n')
                f.write(f'identity: {identity}\n')
                f.write(str(hsp) + '\n')
        else:
            f.write('No significant pairs were found')
    return


def construct_copter(sequences, gaps, reverse=False):
    """ 
    Create linear sequence of the CopTer design, which consists of 3 oligo sequences and 2 gaps between them. The gaps will be represented 
    as "N"s in the sequence. Creates the full sequence from the individual parts and the gap lengths

    Arguements:
        sequences (list or string) -- list or tuple of length 3 with the sequence of each section given as a string
            The sections should ge given as follows: [Extendable, Cooperative, Tertiary]
        gaps (list-like) -- list or tuple of length 2 with the number of nucleotides in each gap given as integers
        reverse (Bool) -- True means the CopTer is a reverse and will be converted to the reverse complement sequence

    Return:
        (str) -- The sequence of the entire copter molecule with N's representing gap nucleotides
    """
    # make sure the sequences and gaps are given as lists or tuples

    if not any([isinstance(sequences, list), isinstance(sequences, tuple)]):
        raise TypeError(f'Sequences must be given as a list or a tuple, type {type(sequences)} was given') 
    
    if not any([isinstance(gaps, list), isinstance(gaps, tuple)]):
        raise TypeError(f'gaps must be given as a list or tuple of integers, type {type(gaps)} was given')
    
    # make sure all the arguments are of the right length
    if len(sequences) != 3:
        raise Exception(f'CopTer designs must contain 3 nucleotide sequences, {len(sequences)} were given')

    if len(gaps) != 2:
        raise Exception(f'The CopTer design cosists of 3 seuqences separated by 2 gaps. {len(gaps)} gap(s) were given')
    
    # make sure the items in the list are the right data types
    for seq in sequences:
        if not isinstance(seq, str):
            raise TypeError(f'The sequences must be strings, type {type(str)} was given')

    for gap in gaps:
        try:
            gap = int(gap)
        except:
            raise TypeError(f'The gaps must all be integer values, type: {type(gap)} was given')

    # construct linearized copter 

    if reverse is True:
        sequence = complement(sequences[2]) + 'N' * gaps[1] + complement(sequences[1]) + 'N' * gaps[0] + complement(sequences[0])

    else:
        sequence = sequences[0] + "N" * gaps[0] + sequences[1] + 'N' * gaps[1] + sequences[2]
    
    return sequence 


def linearize_copter(copter_sequence, gaps=None, reverse=False):
    """ 
    linearize the CopTer molecule sequence with "N"s representing bases in the gap

    Arguments:
        copter_sequence (str) -- the entire sequence of the copter molecule with spacer18s separating each oligo part
        gaps (list) -- list of integers for the gap lengths between the binding regions
        tertiary region, respectively 
        reverse (bool) -- if reverse is True, the reverse complement will be given 
    Return: 
        (str) sequence of entire CopTer molecule with "N"s in place of gap spaces
    """

    copter = clean_sequence(copter_sequence) # returns the copter sequence with '[]' representing the gap regions
    segments = copter.split('[]')
    if reverse is True:
        segments = [reverse_complement(i) for i in segments]
        linear_segments = [segments[1], segments[0], segments[-1]] 
   
    else:
        linear_segments = [segments[-1], segments[0], segments[1]] # put the segments into the right order for a linear view

    copter = '[]'.join(linear_segments)

    number_of_linkers = copter.count('[]')
    if number_of_linkers != 2:
        raise Exception(f'Copters must have 2 linkers, {number_of_linkers} linker(s) found in the sequence provided. Check the sequence and try again')

    if gaps is not None:
        if not any([isinstance(gaps, list), isinstance(gaps, tuple)]):
            raise TypeError(f'gaps must be of type list or tuple, type {type(gaps)} was given')
        if len(gaps) != 2:
            raise Exception('gaps must have two integer values')
        if not all(isinstance(i, int) for i in gaps):
            raise TypeError('gaps must have two integer values')
        linker1 = 'N' * gaps[0]
        linker2 = 'N' * gaps[1]
        copter = copter.replace('[]', linker1, 1)
        copter = copter.replace('[]', linker2)

    return copter


def find_copter_gaps(copter_sequence, reference, reverse=False):
    """ 
    find the appropriate gap lengths for a CopTer sequence. Returns the gap lengths between the binding regions

    Arguments:
        copter_sequence (str) -- CopTer sequence with Spacer18 or Spacer9s separating the binding sequences
        reference (str) -- reference sequence to which the CopTer binds
        reverse (bool) -- if reverse is True, it will be treated as a reverse CopTer and the reverse complement will be used to find the gap lengths 
    Return:
        (list) -- list of gap lengths
    """
    
    reference = clean_sequence(reference)
    copter = linearize_copter(copter_sequence, reverse=reverse)

    segments = copter.split('[]')
    segment1, segment2, segment3 = segments # unpack the segments

    # find where each segment binds to the reference
    seg1_starting = reference.find(segment1)
    seg2_starting = reference.find(segment2)
    seg3_starting = reference.find(segment3)

    if any(pos == -1 for pos in [seg1_starting, seg2_starting, seg3_starting]):
        raise Exception('sequence not found in reference, please check sequences')

    gap1 = seg2_starting - (seg1_starting + len(segment1))
    gap2 = seg3_starting - (seg2_starting + len(segment2))

    gaps = [gap1, gap2]
    print(f'gap lengths: {gaps}')
    return gaps


def linearize_copter_to_reference(copter_sequence, reference):
    """ 
    linearize a copter design by checking where it binds to a reference sequence. No need to know gap lengths or forward or reverse orientation

    """
    reference = clean_sequence(reference)
    copter = clean_sequence(copter_sequence)
    if copter.count('[]') != 2: # make sure it's a copter design written in the right notation

        raise Exception('Does not appear to be a copter format. Check that there are two linkers denoted by "spacer18"')
    segments = copter.split('[]')

    # find binding locations, even if it's a reverse CopTer
    try:
        binding_locations = [binding_location(i, reference) for i in segments]
    except:
        try:
            binding_locations = [binding_location(reverse_complement(i), reference) for i in segments]
            segments = [reverse_complement(i) for i in segments]
        except:
            raise Exception('No binding found for that CopTer sequence and reference. Please check sequences')
    
    # legth of binding_locations should be 3
    if len(binding_locations) != 3:
        raise Exception(f'Sequence Error: Not all binding segments found. {len(binding_locations)} binding locations found.')

    # put the segments and binding loaction into a dataframe
    df = pd.DataFrame({
        'sequences': segments,
        'binding_locations': binding_locations
    })
    # sort them to get the first binding segment on the sequence
    df.sort_values(by='binding_locations', inplace=True)

    df['seq_length'] = df['sequences'].apply(len)
    df['ending_locations'] = df['binding_locations'] + df['seq_length']

    gap_lengths = [] # find the gap lengths
    for i in range(1, 3):
        gap_lengths.append(df.iloc[i]['binding_locations'] - df.iloc[i-1]['ending_locations'])

    # now construct the final copter sequence 
    return construct_copter(list(df['sequences']), gap_lengths)


def linearize_coprimer_to_reference(coprimer, reference):
    """ 
    Linearize a CoPrimer sequence to a reference sequence by finding where the priming and capture regions bind. The CoPrimer must match the 
    template exactly. 

    Arguments:
        coprimer (str) -- sequence of the CoPrimer. The sequences must be in a readable format for clean_sequence()
        reference (str) -- reference sequence or path to FASTA file for the reference sequence
    Return:
        The linearized sequence of the CoPrimer with "N"s representing the Gap bases
    """

    coprimer = clean_sequence(coprimer)
    #verify that the coprimer sequence was cleaned correctly
    if '[]' not in coprimer:
        raise Exception('CoPrimer sequence not cleaned correctly. Check the format')
    
    # process the reference sequence
    if os.path.isfile(reference):
        reference = get_seq_from_fasta(reference)
    elif isinstance(reference, str):
        reference = clean_sequence(reference)
    else:
        raise Exception(f'The reference must be either a path or a string, {type(reference)} given')
    
    # verify that the CoPrimer matches the reference exactly. If not found, try the reverse complement
    primer, capture = coprimer.split('[]')
    
    try:
        primer_binding, reverse = binding_location(primer, reference, return_if_reverse=True)
        capture_binding, reverse = binding_location(capture, reference, return_if_reverse=True)
    except:
            print('CoPrimer not found in the reference sequence')
            return None 


    gap_range = np.arange(2, 20)

    binding_locations = [primer_binding, capture_binding]
    segments = [primer, capture]
    if reverse == True:
        segments = [reverse_complement(i) for i in segments]
    

    df = pd.DataFrame({
        'sequence': segments,
        'binding_location': binding_locations
        }
    )
    df.sort_values(by='binding_location', inplace=True)


    start_gap = df.iloc[0]['binding_location'] + len(df.iloc[0]['sequence'])
    stop_gap = df.iloc[1]['binding_location']
    gap_len = stop_gap - start_gap
    if gap_len not in gap_range:
        raise Exception(f'Gap length of {gap_len} invalid. Check sequences')
    else:
        return df.iloc[0]['sequence'] + 'N' * gap_len + df.iloc[1]['sequence']


def get_fasta(id, email):
    """ 
    Get the FASTA from Entrez given a sequence ID
    Arguments: 
        id (str) -- sequence accession number or other ID number from NCBI
        email (str) -- email to use to fetch FASTA from Entrez
    Return: 
        Seq object 
    """

    handle = Entrez.efetch(db='nucleotide', id=id, rettype='fasta', email=email)
    return SeqIO.read(handle, 'fasta')


def find_binding_template(primer, template_dict):
    """ 
    Find which template in a dictionary a CoPrimer binds to. Returns template name and binding location 
    
    Arguments:
        primer (str) -- primer or CoPrimer sequence
        template_dict (dict) -- dictionary of template sequences to test against the primer
    """
    if not isinstance(template_dict, dict):
        raise Exception('template_dict must be a dictionary')

    for name, seq in template_dict.items():
        if is_exact_match(primer, seq):
            return name, binding_location(primer, seq)
        else:
            continue
    # try again with the reverse complement
    primer = reverse_complement(primer)
    for name, seq in template_dict.items():
        if is_exact_match(primer, seq):
            return name, binding_location(primer, seq)
        else: 
            continue
    return None, None


def combine_records(email, *ids):
    """ 
    Look up the FASTA records of all IDs given, and return a full sequence from all the records. Uses get_fasta
    
    Arguments:
        email (str) -- email address for Entrez query
        ids (str) -- any number of accession or ID numbers for NCBI
        
    Returns:
        SeqIO record of all accession numbers
    """
    all_records = []
    for i in ids:
        try:
            all_records.append(get_fasta(i, email))
        except:
            continue
    full_record = all_records[0]
    for i in all_records[1:]:
        full_record = full_record + i
    
    return full_record


def record_info(path):

    with open(path, 'r+') as f:
        for name, seq in SeqIO.FastaIO.SimpleFastaParser(f):
            return name


def get_description(id, fasta_df):
    """ 
    Use the ID number for a sequence record and get the full description from the FASTA file if the sequence exists in the file

    """
    row = fasta_df[fasta_df['name'].str.contains(id)]
    row.reset_index(inplace=True)
    if len(row) == 0:
        return id
    elif len(row) > 1:
        return row.iloc[0]['name'][0]
    else:
        return row['name'][0]


def sequence_type_differentiation(path, aln=False, fasta=None, out_dir=None, *keywords):
    """ 
    Separate FASTA sequence records based on the presence of key words in the sequence description. If an alignment file is used, the 
    ID number in the alignment file will be used to get the full description from the original FASTA file. The function will print the 
    number of records found for each key word given, and can write separated records for each one in the desired output directory. 

    Arguments:
        path (str, Path-like) -- path to the FASTA or alignment file to use
        aln (bool) -- if True, the ID number from the alignment file will be used to find the full description from the FASTA file. 
            Must also pass the original FASTA file used in the alignment (default False)
        fasta (str, path-like) -- path to the original FASTA file used in the alignment. Only necessary if aln=True. (default False)
        out_dir (str, path-like) -- path to directory where the separated files will be written. No files will be written if None passed. (Default None)
        keywords (str) -- any number of key words to use as identification of subtypes in the sequence data. Not case-sensitive
    """
    # error handling
    if not os.path.isfile(path):
        raise FileExistsError(f'no file found at {path}')

    if aln is True:
        if not os.path.isfile(fasta):
            raise FileExistsError(f'if aln is True, a valid path to the original FASTA file must be given. {fasta} is not a valid path')
        else:
            fasta_df = fasta_to_df(fasta)
    if not out_dir is None:
        # create the directory
        os.makedirs(out_dir, exist_ok=True)

    # create a dataframe from the main input file
    seq_df = fasta_to_df(path)
    seq_df = seq_df.iloc[1:]
    if aln is False:
        seq_df['description'] = seq_df['name']
    else: # if it's an alignment file, we need to find all the full descriptions
        seq_df['description'] = seq_df['name'].apply(lambda x: get_description(x, fasta_df))
        # seq_df['name'] = seq_df['description']
    
    indices_dict = {}

    for word in keywords: 
        word = str(word).strip().lower()
        tmp_df = seq_df[seq_df['description'].str.contains(word, case=False)]
        indices_dict[word] = tmp_df.index


    indices_dict['unclassified'] = [i for i in range(len(seq_df)) if i not in indices_dict.values()]
    
    
    # print number of results per word
    for key, val in indices_dict.items():
        print(f'\n{key}: {len(val)} records found')

    if out_dir:
        for key, val in indices_dict.items():
            if aln is True:
                out_path = os.path.join(out_dir, f'{key}.fasta.aln')
            else:
                out_path = os.path.join(out_dir, f'{key}.fasta')
        tmp_df = seq_df.iloc[val]
        df_to_fasta(tmp_df, out_path)

    return None


def order_format(coprimer_sequence, quencher=None):
    """ 
    Take a CoPrimer sequence and format it so that it's ready for ordering through BioSearch
    Arguments:
        coprimer_sequence (str) -- sequence of the CoPrimer prior to clean_sequence being applied
        quencher (str) -- which quecher to add if necessary. Default is None
    Return: 
        sequence (str) -- formatted CoPrimer sequence

    """

    # check if there is a quencher or dye in the sequence
    labeled = False

    if any([i in coprimer_sequence.upper() for i in dyes_and_quenchers]):
        labeled = True
    
    if quencher != None:
        labeled=True
        quencher = f'[{quencher.strip().upper()}]'

    sequence = clean_sequence(coprimer_sequence)

    linker = "[Spacer 18]aaaaaaaaaaaaaaa[Spacer 18]"
    
    seq_list = sequence.split('[]')
    if len(seq_list) != 2:
        raise Exception('The sequence passed does not appear to be a CoPrimer sequence. CoPrimers are recognized by having Spacer 18 linkers between the binding sections')
    
    if labeled is True:
        seq_list.insert(1, quencher + linker)
        return ''.join(seq_list)
    else:
        return sequence.replace('[]', linker)


def num_records(fasta_path):
    """ 
    Get the number of nucleotide records from a FASTA file

    Arguments:
        fasta_path (path-like) -- path to the FASTA file

    Return:
        (int) -- number of records in the FASTA file
    """

    # error handling 
    if not os.path.isfile(fasta_path):
        raise FileExistsError(f'file path {fasta_path} does not exist. Check path and try again')

    count = 0
    for record in SeqIO.parse(fasta_path, 'fasta'):
        count += 1
    return count


def random_subset_fasta(fasta_path, output_path, number_sequences):
    """ 
    Takes a large FASTA file and writes a new FASTA file that contains a random subset of the larger FASTA file with the number of sequences specified

    Arguments: 
        fasta_path (str, path-like) -- path to the larger FASTA file
        output_path (str, path-like) -- path to write the output subset FASTA file
        number_sequences (int) -- number of sequences to include in the output file. If number_sequences >= the number of records in fasta_path, a randomized version of the 
            FASTA file will be written at output_path and a notification will be displayed. 
    """
    rng = default_rng()

    # -----------------------------error handling------------------------------------------------ 
    if not os.path.isfile(fasta_path):
        raise FileExistsError(f'No file at specified path: {fasta_path}')
    
    # make sure the number of sequences is an integer value
    try:
        number_sequences = int(number_sequences)
    except:
        raise Exception(f'expected int value for number_sequence. Got {type(number_sequences)}')

    # load fasta as a dataframe
    df = fasta_to_df(fasta_path)

    # case if number_sequences >= length of original record
    if len(df) <= number_sequences:
        print(f'number_sequences larger than the original number of records. original: {len(df)}. number_sequences given: {number_sequences}.', 
        'A randomized version of the original data will be written at the output_path location.')
        number_sequences = len(df)

        ints = rng.choice(len(df), size=number_sequences, replace=False)
        df = df.iloc[ints].copy(deep=True)
        df_to_fasta(df, output_path)
        return None
    #-----------------------------------------------------------------------------

    df = rand_subset(df, number_sequences)
    df_to_fasta(df, output_path)
    return None


def mismatch_combinations(sequence, n_mismatches):
    """ 
    Returns a list of all possible sequences with "N" replacing each position. This will allow for matching primer sequences with a reference when there are mismatches
    present in the sequence. 

    Arguments:
        sequence(str) -- sequence of the primer or oligo
        n_mismatches (int) -- number of mismatches allowed (0 to 3)

    Return: 
        iterable of possible sequences with the specified number of mismatches

    """

    if n_mismatches == 0:
        return [sequence]

    if not isinstance(sequence, str):
        raise TypeError(f'sequence must be a string, not {type(sequence)}')
    
    try:
        int(n_mismatches)
    except:
        raise TypeError(f'n_mismatches must be an integer between 0 and 3. {type(n_mismatches)} passed')
    
    if not 0 <= n_mismatches <= 3:
        raise Exception('only mismatches between 0 and 3 allowed')

    seq_l = list(sequence) # convert the sequence to a list
    output = []
    idxs = itt.combinations(range(len(seq_l)), n_mismatches)
    for idx in idxs:
        seq_l2 = seq_l[:] # make a copy of the sequence list
        for i in idx:
            if seq_l2[i] != 'N':
                seq_l2[i] = 'N'
            else:
                continue
            seq = ''.join(seq_l2)
            if seq not in output:
                output.append(seq)

    output.sort(key=lambda x: x.count('N'))
    return output


def dotplot(sequence1, sequence2, word_size=15):
    """ 
    Create a dotplot for comparing sequence similarity between two sequences

    Arguments:
        sequence1 (seq, str) -- first sequence (x-axis)
        sequence2 (seq, str) -- second sequence (y-axis)
        window (int) -- size of window to use

    return -- plotly scatter plot object
    """

    # error handling
    if isinstance(sequence1, Bio.SeqRecord.SeqRecord):
        sequence1 = sequence1.seq
    elif isinstance(sequence1, Bio.Seq.Seq):
        sequence1 = sequence1
    elif isinstance(sequence1, str):
        pass
    else:
        raise TypeError(f'sequence1 must be a Bio.SeqRecord or a string. {type(sequence1)} given')
    
    if isinstance(sequence2, Bio.SeqRecord.SeqRecord):
        sequence2 = sequence2.seq
    elif isinstance(sequence2, Bio.Seq.Seq):
        pass
    elif isinstance(sequence2, str):
        pass
    else:
        raise TypeError(f'sequence2 must be a Bio.SeqRecord or a string. {type(sequence2)} given')
    
    try:
        window = int(word_size)
    except:
        raise TypeError(f'window must be an integer value. {type(word_size)} given')

    # define dictionaries to map sequences to their positions    
    dict_one = {}
    dict_two = {}

    for (seq, section_dict) in [
        (sequence1.upper(), dict_one), 
        (sequence2.upper(), dict_two), 
    ]:
        for i in range(len(seq) - word_size):
            section = seq[i: i + word_size]
            try:
                section_dict[section].append(i)
            except KeyError:
                section_dict[section] = [i]

    # find sub-sequences found in sequence1 and sequence2
    matches = set(dict_one).intersection(dict_two)
    print(f'found {len(matches)} unique matches')

    # separate the lists into x and y to be used in the scatter plot
    x = []
    y = []
    for section in matches: 
        for i in dict_one[section]:
            for j in dict_two[section]:
                x.append(i)
                y.append(j)

    # now create the scatter plot
    df = pd.DataFrame({'sequence1': x, 'sequence2': y})
    fig = px.scatter(df, x='sequence1', y='sequence2', title=f'dotplot similarity (no mismatches allowed), word_size={word_size}')
    fig.update_xaxes(range=[0, len(sequence1)])
    fig.update_traces(marker=dict(size=6))
    return fig, df
  

def find_missing_x(df):
    """ Get a list of the missing x values in the dotplot output
    Arguments:
        df (Pandas DataFrame) -- dataframe output from dotplot
    Return: list of all missing x values in the range of the sequence index
    """
    vals = df['sequence1'].values
    return sorted([x for x in tqdm(range(vals.min(), vals.max())) if x not in vals])


def find_missing_y(df):
    """ Get a list of the missing y values in the dotplot output
    Arguments:
        df (Pandas DataFrame) -- dataframe output from dotplot
    Return: list of all missing y values in the range of the sequence index
    """
    vals = df['sequence2'].values
    return sorted([y for y in tqdm(range(vals.min(), vals.max())) if y not in vals])


def largest_gap(lst):
    """ 
    Determine the largest gap in either the x or y data from a dotplot. 

    Arguments: 
        lst (list, iterable) -- iterable of the missing x or y values from dotplot. Output from find_missing_x or find_missing_y

    Return: (list) - list of the largest consecutive gap in the data given. If multiple, a list of lists will be returned 
    """
    lst = sorted(lst)
    multiple_longest = [] # if multiple ranges have the same length, give all of them
    longest = []
    counter = 0
    for i in range(len(lst)-1):
        if lst[i+1] == lst[i] + 1:
            counter += 1
        else:
            range_i = lst[i- counter: i]
            counter = 0 # reset the counter
            if len(range_i) > len(longest):
                longest = range_i
                multiple_longest = [] # clear the multiple_longest list if a new longest sequence is found
            elif len(range_i) == len(longest):
                if longest not in multiple_longest:
                    multiple_longest.append(longest)
                multiple_longest.append(range_i)
            else:
                continue
    if len(longest) == 0:
        return lst
    
    if len(multiple_longest) > 0:
        return multiple_longest
    else:
        return longest


def standard_primers(template, tm):
    """ 
    Design standard PCR primers from a template sequence. Returns forward and reverse primer in Dictionary

    Arguments:
        template (str, seq) -- template sequence to use in designing the primers. Template should be desired amplicon +1 on each end
        tm (int) -- Desired TM of the primers
    Return:
        forward (dict)
        reverse (dict)
    """
    if isinstance(template, Bio.SeqRecord.SeqRecord):
        template = str(template.seq)
    
    if not isinstance(tm, int) and not isinstance(tm, float):
        try:
            tm = int(tm)
        except:
            raise Exception(f'TM must be a numerical type (int or float). {type(tm)} given')

    forward = trim_to_tm(template[1:], tm, method='right', return_tm=True)
    reverse = trim_to_tm(template[:-1], tm, method='left', return_tm=True)
    reverse['sequence'] = reverse_complement(reverse['sequence'])
    return forward, reverse


def make_probe(template, tm, forward, reverse, separation=5):
    """ 
    Design a probe sequence for a standard primer assay. 

    Arguments: 
        template(str, seq) -- template sequence for the assay
        tm (num) -- tm of the probe
        forward (str, seq) -- sequence of the forward primer
        reverse (str, seq) -- sequence of the reverse primer
        separation (int) -- Basepair separation between the end of the forward primer and the start of the probe (Default=5)
    Return 
        Dict of probe sequence and TM

    """
    if isinstance(template, Bio.SeqRecord.SeqRecord):
        template = str(template.seq)

    # get correct type for forward
    if isinstance(forward, dict):
        forward = forward['sequence']
    elif isinstance(forward, Bio.SeqRecord.SeqRecord):
        forward = str(forward.seq)
    elif isinstance(forward, str):
        pass

    # get correct type for reverse
    if isinstance(reverse, dict):
        reverse = reverse['sequence']
    elif isinstance(reverse, Bio.SeqRecord.SeqRecord):
        reverse = str(reverse.seq)
    elif isinstance(reverse, str):
        pass

    # check that tm is numerical type
    if not isinstance(tm, int) and not isinstance(tm, float):
        try:
            tm = int(tm)
        except:
            raise Exception(f'TM must be a numerical type (int or float). {type(tm)} given')
        
    reverse = reverse_complement(reverse) #convert reverse back to match template

    forward_binding = binding_location(forward, template, method='span')
    reverse_binding = binding_location(reverse, template, method='span')

    probe = trim_to_tm(template[forward_binding[-1] + separation:], tm, method='right', return_tm=True)
    counter = 1
    while probe['sequence'][0] == 'G':
        probe = trim_to_tm(template[forward_binding[-1] + separation + counter:], tm, method='right', return_tm=True)
        counter += 1

    # check that it doesn't overlap the binding for the reverse
    probe_binding = binding_location(probe['sequence'], template, method='span')
    if probe_binding[-1] >= reverse_binding[0]:
        raise Exception('the probe overlaps with the reverse primer binding. Adjust separation value or TM to fix the issue')
    else:
        return probe


def risk_categories(aln_path, primer_sequence, number_mismatches=[0, 1, 2]):
    """ 
    categorize sequences according to risk based on the number of mismatches on the primer sequence. The
    function will return the list of accession numbers associated with each mismatch number. eg. a list 
    with 0 mismatches, a list of sequences with 1 mismatch, and a list of sequences with 2+ mismatches. 

    Arguments: 
        aln_path (str, path-like) -- path to the alignment file
        primer_sequence (str) -- sequence of the primer or CoPrimer (must be linearized if CoPrimer)
        number_mismatches (iterable) -- mismatch categorizations. This will dictate the number of returns from the function. 
            maximum of 3 mismatches can be given
    
    Return: 
        list(s) of ID numbers for sequences of the specified mismatch categories. Default will return 3 lists sorted from lowest mismatch count to highest
    """

    # error handling
    if not os.path.exists(aln_path):
        raise FileNotFoundError('the alignment file cannot be found')
    
    primer_sequence = clean_sequence(primer_sequence, allow_ambiguous=True)
    if '[]' in primer_sequence:
        raise Exception('The CoPrimer must be linearized with "N" denoting gap bases')
    
    #check if iterable
    try:
        iter(number_mismatches)
    except:
        raise TypeError(f'number_mismatches must be an iterable with number of mismatches such as list or tuple. {type(number_mismatches)} given')
    
    try:
        number_mismatches = sorted([int(num) for num in number_mismatches])
    except:
        raise TypeError('Each value in number_mismatches must be integer values.')
    
    if max(number_mismatches) > 3:
        raise Exception(f'Cannot assess binding with mismatch > 3. A maximum of {max(number_mismatches)} was given')

    # first, build a sequence matrix from the alignment file 
    print('building sequence matrix...')
    matrix = seq_matrix(aln_path)

    # get the IDs from the alignment file
    df = fasta_to_df(aln_path)
    IDs = df['name']
    

    # cut down the matrix to greatly reduce computation time. First step is to find the binding region
    # find a sequence that binds to the primer to get the binding region
    for i in range(matrix.shape[0]):
        seq = str(''.join(matrix[i,::]))
        if i == matrix.shape[0]:
            raise Exception('No binding found for any of the sequences')
        try:
            binding_loc = binding_location(primer_sequence, seq, 0, 'span')
            break
        except:
            pass
    
    matrix = matrix[:,binding_loc[0]-5:binding_loc[1]+5] # added 5 bases on either side of the binding region

    # find sequences that aren't successfully sequenced in that region and drop them
    # non-sequenced region is denoted by '-'
    drop_indices = []
    for i in range(matrix.shape[0]):
        seq = str(''.join(matrix[i,::]))
        if '--' in seq:
            drop_indices.append(i)
        
    IDs.drop(index=drop_indices, inplace=True)
    # IDs = IDs.reset_index()
    # IDs.squeeze()

    matrix = np.delete(matrix, drop_indices, axis=0) # drop the same indices from the sequence matrix 

    final_lists = [] # list of lists to return with the corresponding number of mismatches
    
    for mismatch in number_mismatches:
        mismatch_list = [] # list of sequences that match the current mismatch count
        for i in range(matrix.shape[0]):
            seq = str(''.join(matrix[i,::]))
            if "NN" in seq:
                continue
            try:
                binding_location(primer_sequence, seq, mismatches=mismatch)
            except:
                continue
            mismatch_list.append(IDs.iloc[i].replace('>', ''))
        mismatch_list = [id for id in mismatch_list if id not in flatten(final_lists)]
        final_lists.append(mismatch_list)
    return final_lists

            
def check_binding(primer_sequence, ID, email, binding_region=None):
    """
    Check the binding of a primer with a sequence (by ID number)

    Arguments: 
        primer_sequence (str) -- Sequence of the primer or CoPrimer (formatted and linearized already)
        ID (str) -- ID number of sequence to acquire from NCBI database
        email (str) -- email to use to access NCBI
        binding_regions (tuple) -- binding_region on the sequence if known. If None, a binding region will be searched using binding_location function
    Return:
        None. The binding of the primer will be printed to the console
    """

    seq = str(get_fasta(ID, email).seq)
    region_error = 'binding region should be a tuple of 2 values'
    if binding_region != None:
        if type(binding_region) not in ['list', 'tuple']:
            raise Exception(region_error)
        if len(binding_region) != 2:
            raise Exception(region_error)
    else:
        try:
            binding_region = binding_location(primer_sequence, seq, mismatches=3, method='span')
            seq = seq[binding_region[0]-5: binding_region[1]+5]
            bind_primer(primer_sequence, seq)
        except:
            bind_primer(primer_sequence, seq)
    return None


def write_multi_fasta(id_list, path, email):
    """ 
    Get the sequences from a list of accession numbers, and write them as a multi-FASTA file

    Arguments: 
        id_list (list) -- list of ID or accession numbers from NCBI
        path (str, pathlike) -- path where the multi-fasta will be written
        email (str) -- email to be passed to NCBI
    Return:
        None
    """
    if len(id_list) == 0:
        print('No ID numbers in list')
        return None

    parent = parent_dir(path)
    os.makedirs(parent, exist_ok=True)

    seqs = []
    for id in id_list: 
        seqs.append(get_fasta(id, email))

    SeqIO.write(seqs, path, 'fasta')
    return None


def split_template(template_sequence, tm):
    """ 
    Split the DNA template into two overlapping segments that can be extended in PCR reaction to create the full DNA template. 
    This is to mitigate the risk of whole amplicon contamination
    Arguments:
        template_sequence (str) -- sequence of the whole template 
        tm (int) -- melting temperature of the overlapping region
    output: 
        (tuple) -- top and bottom strands of the DNA template with an overlapping region for Taq extension (top_strand, bottom_strand, overlap_TM)
    """

    #error handling 
    if not isinstance(template_sequence, str):
        raise TypeError(f'template sequence must be a string, {type(template_sequence)} given')
    try:
        tm = int(tm)
    except:
        raise TypeError(f'tm must be an integer value, {type(tm)} given')
    
    template_sequence = clean_sequence(template_sequence, True)
    # split the template in half 
    top = template_sequence[:len(template_sequence)//2]
    bottom = complement(template_sequence[len(template_sequence)//2:])

    # create the overlapping region
    overlap = trim_to_tm(top, tm, method='left', return_tm=True)
    overlap_tm = overlap['tm']
    overlap_c = complement(overlap['sequence']) # make sure to use the complement on the bottom strand

    # concatenate the overlapping region to the bottom strand
    bottom = overlap_c + bottom
    print(overlap)

    return (top, bottom, round(overlap_tm, 2))
